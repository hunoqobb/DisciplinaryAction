from PyQt5.QtWidgets import (QWidget, QVBoxLayout, QHBoxLayout, QFormLayout,
                            QLabel, QLineEdit, QComboBox, QPushButton, QTableWidget,
                            QTableWidgetItem, QHeaderView, QMessageBox, QDialog,
                            QFileDialog)
from ui.help_dialogs import HelpDialog
from PyQt5.QtCore import Qt, QDate
from utils.data_manager import DataManager
import openpyxl
from openpyxl.styles import Font, Alignment
from datetime import datetime

class StatisticsTab(QWidget):
    def __init__(self, db):
        super().__init__()
        self.db = db
        self.initUI()
        self.loadData()
    
    def initUI(self):
        main_layout = QVBoxLayout(self)
        
        # 创建查询表单
        form_layout = QFormLayout()
        
        # 学生信息
        self.name_edit = QLineEdit()
        form_layout.addRow("姓名:", self.name_edit)
        
        self.grade_combo = QComboBox()
        form_layout.addRow("年级:", self.grade_combo)
        
        self.class_combo = QComboBox()
        form_layout.addRow("专业班级:", self.class_combo)
        
        self.punishment_type_combo = QComboBox()
        self.punishment_type_combo.addItem("全部")
        self.punishment_type_combo.addItems(["警告", "严重警告", "记过", "留校察看一学期", "留校察看一学年", "留校察看两学年", "开除学籍"])
        form_layout.addRow("处分类型:", self.punishment_type_combo)
        
        # 查询按钮
        button_layout = QHBoxLayout()
        self.search_btn = QPushButton("查询")
        self.search_btn.clicked.connect(self.searchRecords)
        button_layout.addWidget(self.search_btn)
        
        self.delete_btn = QPushButton("删除")
        self.delete_btn.clicked.connect(self.deleteStudent)
        button_layout.addWidget(self.delete_btn)
        
        self.export_btn = QPushButton("导出Excel")
        self.export_btn.clicked.connect(self.exportToExcel)
        button_layout.addWidget(self.export_btn)
        
        # 表格
        self.table = QTableWidget()
        self.table.setColumnCount(8)
        self.table.setHorizontalHeaderLabels(["姓名", "性别", "年级", "专业班级", "核销所需积分", "已获得积分", "尚需积分", "状态"])
        self.table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
        self.table.setEditTriggers(QTableWidget.NoEditTriggers)
        self.table.setSelectionBehavior(QTableWidget.SelectRows)
        self.table.doubleClicked.connect(self.showStudentDetail)
        
        # 添加表头点击事件
        self.table.horizontalHeader().sectionClicked.connect(self.onHeaderClicked)
        self.current_sort_column = -1  # 当前排序列
        self.sort_order = 0  # 0: 初始状态, 1: 升序, 2: 降序
        
        # 添加到主布局
        main_layout.addLayout(form_layout)
        main_layout.addLayout(button_layout)
        main_layout.addWidget(self.table)
    
    def loadData(self):
        try:
            # 加载年级数据
            grades = DataManager.load_grades()
            self.grade_combo.clear()
            self.grade_combo.addItem("所有", None)
            for grade_id, grade_name in grades:
                self.grade_combo.addItem(grade_name, grade_id)
            
            # 加载专业班级数据
            classes = DataManager.load_classes()
            self.class_combo.clear()
            self.class_combo.addItem("所有", None)
            for class_id, _, major, _ in classes:
                self.class_combo.addItem(major, class_id)
        except Exception as e:
            print(f"加载数据失败：{str(e)}")
    
    def searchRecords(self):
        try:
            # 构建查询条件
            conditions = []
            params = []
            
            name = self.name_edit.text().strip()
            if name:
                conditions.append("s.name LIKE ?")
                params.append(f"%{name}%")
            
            grade_id = self.grade_combo.currentData()
            if grade_id:
                conditions.append("s.grade_id = ?")
                params.append(grade_id)
            
            class_id = self.class_combo.currentData()
            if class_id:
                conditions.append("s.class_id = ?")
                params.append(class_id)
            
            punishment_type = self.punishment_type_combo.currentText()
            if punishment_type != "全部":
                conditions.append("pt.name = ?")
                params.append(punishment_type)
            
            # 构建SQL查询
            where_clause = " AND ".join(conditions) if conditions else "1=1"
            
            query = f"""
                SELECT s.id as student_id, s.name, s.gender, s.grade_id, s.class_id,
                       SUM(p.required_points) as total_required_points,
                       COALESCE((SELECT SUM(points) FROM activities WHERE student_id = s.id), 0) as earned_points,
                       CASE WHEN COUNT(*) = SUM(CASE WHEN p.is_cleared = 1 THEN 1 ELSE 0 END) THEN '已核销' ELSE '' END as status
                FROM students s
                JOIN punishments p ON s.id = p.student_id
                JOIN punishment_types pt ON p.type_id = pt.id
                WHERE {where_clause}
                GROUP BY s.id, s.name, s.gender, s.grade_id, s.class_id
                ORDER BY s.name
            """
            
            self.db.cursor.execute(query, params)
            records = self.db.cursor.fetchall()
            
            # 加载年级和专业班级数据，用于显示名称
            grades_dict = {grade_id: grade_name for grade_id, grade_name in DataManager.load_grades()}
            classes_dict = {class_id: major for class_id, _, major, _ in DataManager.load_classes()}
            
            self.table.setRowCount(len(records))
            for row, record in enumerate(records):
                student_id, name, gender, grade_id, class_id, required, earned, status = record
                remaining = max(0, required - earned)
                
                # 根据ID获取年级和专业班级名称
                grade_name = grades_dict.get(grade_id, "未知")
                class_name = classes_dict.get(class_id, "未知")
                
                # 创建表格项并存储学生ID
                name_item = QTableWidgetItem(name)
                name_item.setData(Qt.UserRole, student_id)  # 存储学生ID
                
                self.table.setItem(row, 0, name_item)
                self.table.setItem(row, 1, QTableWidgetItem(gender))
                self.table.setItem(row, 2, QTableWidgetItem(grade_name))
                self.table.setItem(row, 3, QTableWidgetItem(class_name))
                self.table.setItem(row, 4, QTableWidgetItem(str(required)))
                self.table.setItem(row, 5, QTableWidgetItem(str(earned)))
                self.table.setItem(row, 6, QTableWidgetItem(str(remaining)))
                self.table.setItem(row, 7, QTableWidgetItem(status))
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"查询失败：{str(e)}")
    
    def showStudentDetail(self):
        try:
            row = self.table.currentRow()
            if row < 0:
                return
            
            # 获取学生信息，包括存储在表格项中的学生ID
            name_item = self.table.item(row, 0)
            student_name = name_item.text()
            student_id = name_item.data(Qt.UserRole)  # 获取存储的学生ID
            grade_name = self.table.item(row, 2).text()
            class_name = self.table.item(row, 3).text()
            
            # 创建详情对话框
            dialog = QDialog(self)
            dialog.setWindowFlags(dialog.windowFlags() & ~Qt.WindowContextHelpButtonHint)
            dialog.setWindowTitle(f"{student_name} - 详细信息")
            dialog.setMinimumWidth(600)
            
            layout = QVBoxLayout(dialog)
            
            # 处分记录表格
            punishment_table = QTableWidget()
            punishment_table.setColumnCount(5)
            punishment_table.setHorizontalHeaderLabels(["处分类型", "处分原因", "处分日期", "核销所需积分", "状态"])
            punishment_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            punishment_table.setEditTriggers(QTableWidget.NoEditTriggers)
            punishment_table.setSelectionBehavior(QTableWidget.SelectRows)
            punishment_table.doubleClicked.connect(lambda: self.handlePunishmentDoubleClick(student_id))
            punishment_table.horizontalHeader().setToolTip("")  # 移除帮助提示
            # 启用表格排序功能
            punishment_table.setSortingEnabled(True)
            
            # 查询处分记录 - 直接使用学生ID查询，避免同名问题
            self.db.cursor.execute("""
                SELECT pt.name, p.reason, p.date, p.required_points,
                       CASE WHEN p.is_cleared = 1 THEN '已核销' ELSE '未核销' END,
                       p.id, s.id
                FROM punishments p
                JOIN punishment_types pt ON p.type_id = pt.id
                JOIN students s ON p.student_id = s.id
                WHERE s.id = ?
                ORDER BY p.date DESC
            """, (student_id,))
            punishments = self.db.cursor.fetchall()
            
            punishment_table.setRowCount(len(punishments))
            for row, record in enumerate(punishments):
                for col in range(5):  # 只显示前5列
                    item = QTableWidgetItem(str(record[col]))
                    # 设置适当的数据类型以便正确排序
                    if col == 3:  # 核销所需积分列
                        item.setData(Qt.DisplayRole, int(record[col]))
                    punishment_table.setItem(row, col, item)
                # 存储处分ID和学生ID作为表格项的数据
                punishment_table.item(row, 0).setData(Qt.UserRole, (record[5], record[6]))
            
            # 活动记录表格
            activity_table = QTableWidget()
            activity_table.setColumnCount(4)
            activity_table.setHorizontalHeaderLabels(["活动内容", "活动日期", "活动时长", "获得积分"])
            activity_table.horizontalHeader().setSectionResizeMode(QHeaderView.Stretch)
            activity_table.horizontalHeader().setToolTip("")  # 移除帮助提示
            # 启用表格排序功能
            activity_table.setSortingEnabled(True)
            
            # 查询活动记录 - 直接使用学生ID查询，避免同名问题
            self.db.cursor.execute("""
                SELECT content, date, duration, points
                FROM activities a
                JOIN students s ON a.student_id = s.id
                WHERE s.id = ?
                ORDER BY date DESC
            """, (student_id,))
            activities = self.db.cursor.fetchall()
            
            activity_table.setRowCount(len(activities))
            for row, record in enumerate(activities):
                for col, value in enumerate(record):
                    item = QTableWidgetItem(str(value))
                    # 设置适当的数据类型以便正确排序
                    if col == 1:  # 日期列
                        item.setData(Qt.DisplayRole, QDate.fromString(str(value), "yyyy-MM-dd"))
                    elif col == 2:  # 时长列
                        item.setData(Qt.DisplayRole, float(value))
                    elif col == 3:  # 积分列
                        item.setData(Qt.DisplayRole, int(value))
                    activity_table.setItem(row, col, item)
            
            # 添加表格到对话框
            layout.addWidget(QLabel("处分记录"))
            layout.addWidget(punishment_table)
            layout.addWidget(QLabel("公益活动记录"))
            layout.addWidget(activity_table)
            
            # 显示对话框
            dialog.exec_()
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"加载详细信息失败：{str(e)}")

    def handlePunishmentDoubleClick(self, student_id):
        try:
            # 获取处分表格（即发送信号的对象）
            punishment_table = self.sender()
            if not punishment_table:
                return
                
            # 获取当前选中的处分行
            punishment_row = punishment_table.currentRow()
            if punishment_row < 0:
                return
                
            # 获取当前选中的处分记录
            punishment_row = self.sender().currentRow()
            punishment_item = self.sender().item(punishment_row, 0)
            if not punishment_item:
                return
                
            # 获取处分ID和学生ID
            punishment_id, _ = punishment_item.data(Qt.UserRole)  # 学生ID已经通过参数传入
            
            # 检查处分状态
            status_item = self.sender().item(punishment_row, 4)
            if status_item.text() == "已核销":
                QMessageBox.information(self, "提示", "该处分已经核销")
                return
            
            # 弹出确认对话框
            reply = QMessageBox.question(self, "确认", "是否要核销该处分？",
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            
            if reply == QMessageBox.No:
                return
            
            # 检查公益积分是否足够
            self.db.cursor.execute("""
                SELECT 
                    (SELECT COALESCE(SUM(points), 0) FROM activities WHERE student_id = ?) as earned_points,
                    (SELECT required_points FROM punishments WHERE id = ?) as required_points
            """, (student_id, punishment_id))
            
            points = self.db.cursor.fetchone()
            if not points:
                QMessageBox.warning(self, "错误", "无法获取积分信息")
                return
                
            earned_points, required_points = points
            
            if earned_points < required_points:
                QMessageBox.warning(self, "提示", f"公益积分不足，还需要{required_points - earned_points}分")
                return
            
            # 只更新处分状态为已核销，不扣除活动积分
            self.db.cursor.execute("""
                UPDATE punishments
                SET is_cleared = 1
                WHERE id = ?
            """, (punishment_id,))
            
            self.db.conn.commit()
            
            # 更新表格显示
            status_item.setText("已核销")
            self.searchRecords()  # 刷新主表格
            
            QMessageBox.information(self, "成功", "处分已成功核销")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"核销处分失败：{str(e)}")
            self.db.conn.rollback()
    


    def onHeaderClicked(self, logical_index):
        if self.current_sort_column == logical_index:
            # 同一列的点击，切换排序状态
            self.sort_order = (self.sort_order + 1) % 3
        else:
            # 不同列的点击，重置为升序
            self.current_sort_column = logical_index
            self.sort_order = 1

        if self.sort_order == 0:
            # 恢复初始排序（按姓名排序）
            self.table.sortItems(0)
        else:
            # 根据点击的列进行排序
            self.table.sortItems(logical_index, Qt.AscendingOrder if self.sort_order == 1 else Qt.DescendingOrder)
    
    def exportToExcel(self):
        try:
            # 获取保存文件路径
            file_name = f"学生处分核销统计_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            save_path, _ = QFileDialog.getSaveFileName(self, "保存Excel文件", file_name, "Excel文件 (*.xlsx)")
            
            if not save_path:
                return
            
            # 创建Excel工作簿
            wb = openpyxl.Workbook()
            
            # 创建学生统计表
            ws_stats = wb.active
            ws_stats.title = "学生统计"
            
            # 设置表头
            headers = ["姓名", "性别", "年级", "专业班级", "核销所需积分", "已获得积分", "尚需积分", "状态"]
            for col, header in enumerate(headers, 1):
                cell = ws_stats.cell(row=1, column=col)
                cell.value = header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
            
            # 写入学生统计数据
            for row in range(self.table.rowCount()):
                for col in range(self.table.columnCount()):
                    cell = ws_stats.cell(row=row+2, column=col+1)
                    cell.value = self.table.item(row, col).text()
                    cell.alignment = Alignment(horizontal='center')
            
            # 创建处分记录表
            ws_punishments = wb.create_sheet("处分记录")
            ws_punishments.append(["姓名", "处分类型", "处分原因", "处分日期", "核销所需积分", "状态"])
            
            # 查询处分记录
            self.db.cursor.execute("""
                SELECT s.name, pt.name, p.reason, p.date, p.required_points,
                       CASE WHEN p.is_cleared = 1 THEN '已核销' ELSE '未核销' END
                FROM punishments p
                JOIN students s ON p.student_id = s.id
                JOIN punishment_types pt ON p.type_id = pt.id
                ORDER BY s.name, p.date DESC
            """)
            
            for record in self.db.cursor.fetchall():
                ws_punishments.append(record)
            
            # 创建活动记录表
            ws_activities = wb.create_sheet("公益活动记录")
            ws_activities.append(["姓名", "活动内容", "活动日期", "活动时长", "获得积分"])
            
            # 查询活动记录
            self.db.cursor.execute("""
                SELECT s.name, a.content, a.date, a.duration, a.points
                FROM activities a
                JOIN students s ON a.student_id = s.id
                ORDER BY s.name, a.date DESC
            """)
            
            for record in self.db.cursor.fetchall():
                ws_activities.append(record)
            
            # 调整列宽
            for ws in [ws_stats, ws_punishments, ws_activities]:
                for column in ws.columns:
                    max_length = 0
                    column = list(column)
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    adjusted_width = (max_length + 2)
                    ws.column_dimensions[column[0].column_letter].width = adjusted_width
            
            # 保存文件
            wb.save(save_path)
            QMessageBox.information(self, "成功", "Excel文件导出成功！")
            
        except Exception as e:
            QMessageBox.critical(self, "错误", f"导出Excel失败：{str(e)}")

    def deleteStudent(self):
        try:
            # 获取当前选中的行
            row = self.table.currentRow()
            if row < 0:
                QMessageBox.warning(self, "错误", "请选择要删除的学生")
                return
            
            # 获取学生姓名
            student_name = self.table.item(row, 0).text()
            
            # 确认删除
            reply = QMessageBox.question(self, "确认", f"确定要删除学生 {student_name} 的所有记录吗？\n此操作将删除该学生的所有处分记录和公益活动记录。",
                                       QMessageBox.Yes | QMessageBox.No, QMessageBox.No)
            
            if reply == QMessageBox.No:
                return
            
            # 获取学生ID
            self.db.cursor.execute("SELECT id FROM students WHERE name = ?", (student_name,))
            student = self.db.cursor.fetchone()
            if not student:
                QMessageBox.warning(self, "错误", "找不到该学生的记录")
                return
            
            student_id = student[0]
            
            # 删除该学生的所有记录
            self.db.cursor.execute("DELETE FROM activities WHERE student_id = ?", (student_id,))
            self.db.cursor.execute("DELETE FROM punishments WHERE student_id = ?", (student_id,))
            self.db.cursor.execute("DELETE FROM students WHERE id = ?", (student_id,))
            
            self.db.conn.commit()
            
            # 从表格中移除该行
            self.table.removeRow(row)
            
            QMessageBox.information(self, "成功", f"已删除学生 {student_name} 的所有记录")
            
        except Exception as e:
            self.db.conn.rollback()
            QMessageBox.critical(self, "错误", f"删除记录失败：{str(e)}")